"""
Pantalla de importación masiva de contribuyentes desde plantilla Excel.
"""
import io
import tempfile
from pathlib import Path

import streamlit as st

from iibb.db import get_session
from iibb.exporters.plantilla_contribuyentes import generar_plantilla
from iibb.importers.contribuyentes_excel import importar_contribuyentes_excel


def render_importar_contribuyentes():
    col_back, col_title = st.columns([1, 5])
    with col_back:
        if st.button("← Volver"):
            st.session_state["pantalla"] = "dashboard"
            st.rerun()
    with col_title:
        st.title("Importar contribuyentes desde Excel")

    st.markdown("---")

    # ── Descarga de plantilla ─────────────────────────────────────────────────
    st.subheader("1. Descargá la plantilla")
    st.markdown(
        "La plantilla tiene **5 hojas**: Contribuyentes, Jurisdicciones, Actividades, "
        "Coeficientes y Alícuotas. Completá cada hoja siguiendo las instrucciones "
        "incluidas en la fila 2 (fondo amarillo)."
    )

    if st.button("⬇️ Descargar plantilla Excel", type="primary"):
        with tempfile.TemporaryDirectory() as tmp:
            destino = Path(tmp) / "plantilla_contribuyentes_iibb.xlsx"
            generar_plantilla(destino)
            data = destino.read_bytes()
        st.download_button(
            label="📥 Guardar plantilla_contribuyentes_iibb.xlsx",
            data=data,
            file_name="plantilla_contribuyentes_iibb.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    st.markdown("---")

    # ── Carga del archivo ─────────────────────────────────────────────────────
    st.subheader("2. Cargá el archivo completado")

    archivo = st.file_uploader(
        "Seleccioná el Excel con los contribuyentes",
        type=["xlsx"],
        help="Solo se acepta el formato .xlsx (Excel moderno).",
    )

    if archivo is None:
        st.info("Esperando el archivo Excel...")
        return

    # Vista previa mínima ─ muestra cuántas filas hay en hoja 1
    import openpyxl  # noqa: PLC0415
    try:
        wb_prev = openpyxl.load_workbook(io.BytesIO(archivo.read()), data_only=True)
        archivo.seek(0)
        ws_prev = next(
            (wb_prev[n] for n in wb_prev.sheetnames if "contribuyente" in n.lower()),
            None,
        )
        if ws_prev:
            filas = sum(
                1 for row in ws_prev.iter_rows(min_row=3, values_only=True)
                if any(c is not None and str(c).strip() != "" for c in row)
            )
            st.success(f"Archivo válido · **{filas}** contribuyente(s) detectado(s) en la hoja 1.")
        else:
            st.warning("No se encontró la hoja de Contribuyentes en el archivo.")
    except Exception as e:
        st.error(f"No se pudo leer el archivo: {e}")
        return

    st.markdown("---")

    # ── Botón de importación ──────────────────────────────────────────────────
    st.subheader("3. Importar")
    st.markdown(
        "- Si el CUIT **no existe** → se crea el contribuyente con todos sus datos.\n"
        "- Si el CUIT **ya existe** → se actualizan sus datos y se agregan "
        "jurisdicciones, actividades y coeficientes faltantes."
    )

    if st.button("🚀 Importar contribuyentes", type="primary", use_container_width=True):
        with tempfile.TemporaryDirectory() as tmp:
            ruta = Path(tmp) / archivo.name
            ruta.write_bytes(archivo.getvalue())

            session = get_session()
            try:
                with st.spinner("Procesando..."):
                    resultado = importar_contribuyentes_excel(ruta, session)
                session.commit()
            except Exception as e:
                session.rollback()
                st.error(f"Error durante la importación: {e}")
                return
            finally:
                session.close()

        # Mostrar resultado
        col_n, col_u, col_e = st.columns(3)
        col_n.metric("Nuevos", resultado.contribuyentes_nuevos)
        col_u.metric("Actualizados", resultado.contribuyentes_actualizados)
        col_e.metric("Errores", len(resultado.errores))

        if resultado.errores:
            st.error("**Errores detectados:**")
            for err in resultado.errores:
                st.markdown(f"- {err}")

        if resultado.advertencias:
            st.warning("**Advertencias:**")
            for adv in resultado.advertencias:
                st.markdown(f"- {adv}")

        total = resultado.contribuyentes_nuevos + resultado.contribuyentes_actualizados
        if total > 0:
            st.success(
                f"✅ Importación completada: {resultado.contribuyentes_nuevos} nuevo(s), "
                f"{resultado.contribuyentes_actualizados} actualizado(s)."
            )
            if st.button("Ver contribuyentes"):
                st.session_state["pantalla"] = "dashboard"
                st.rerun()
