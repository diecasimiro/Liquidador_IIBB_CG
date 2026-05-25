"""
Genera la plantilla Excel para importación masiva de contribuyentes.
5 hojas: Contribuyentes, Jurisdicciones, Actividades, Coeficientes, Alicuotas.
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_AZUL     = "1F4E79"
_AZUL_CL  = "BDD7EE"
_VERDE    = "375623"
_GRIS     = "F2F2F2"
_AMARILLO = "FFF2CC"


def _header(ws, row, col, text, bg=_AZUL, fg="FFFFFF"):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    return c


def _instruccion(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = PatternFill("solid", fgColor=_AMARILLO)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.font = Font(italic=True, size=9)
    return c


def generar_plantilla(destino: Path | None = None) -> Path:
    if destino is None:
        downloads = Path.home() / "Downloads"
        downloads.mkdir(exist_ok=True)
        destino = downloads / "plantilla_contribuyentes_iibb.xlsx"

    wb = openpyxl.Workbook()

    _hoja_contribuyentes(wb)
    _hoja_jurisdicciones(wb)
    _hoja_actividades(wb)
    _hoja_coeficientes(wb)
    _hoja_alicuotas(wb)
    _hoja_ayuda(wb)

    # Quitar hoja por defecto
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(destino)
    return destino


# ─────────────────────────────────────────────────────────────────────────────

def _hoja_contribuyentes(wb):
    ws = wb.create_sheet("1_Contribuyentes")
    ws.sheet_properties.tabColor = "1F4E79"

    columnas = [
        ("CUIT *",               "Ej: 30-71223438-1  (con guiones)",            18),
        ("Razón Social *",       "Ej: EMPRESA EJEMPLO S.A.",                     35),
        ("N° Inscripción CM",    "Ej: 901-663948-1",                             20),
        ("Jurisdicción Sede *",  "Código 3 dígitos: 901=CABA, 902=BsAs...",      18),
        ("Tipo Contribuyente",   "Resto / Empresa unipersonal / etc.",            22),
        ("Naturaleza Jurídica",  "Sociedad Anónima / SRL / etc.",                 25),
        ("Cierre Ejercicio Mes", "Número: 6=Junio, 12=Diciembre",                20),
        ("Cierre Ejercicio Día", "Número: 30 o 31",                              18),
        ("Domicilio Fiscal",     "Dirección completa",                            40),
        ("Domicilio Actividades","Si difiere del fiscal",                         40),
        ("Notas",                "Notas internas opcionales",                     30),
    ]

    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 40

    for i, (col_name, instruccion, ancho) in enumerate(columnas, 1):
        _header(ws, 1, i, col_name)
        _instruccion(ws, 2, i, instruccion)
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # 3 filas de ejemplo
    ejemplos = [
        ["30-71223438-1", "AMERICAN IMPLANT S.A.", "901-663948-1", "901",
         "Resto", "Sociedad Anónima", 6, 30,
         "ALVAREZ JONTE AV. 1647 P.11 OF.F - CABA", "Cerrito 1079 - ITUZAINGO", ""],
        ["20-12345678-9", "GARCIA MARIA LAURA", "", "902",
         "Empresa unipersonal", "Empresa Unipersonal", 12, 31,
         "SAN MARTIN 456 - ROSARIO", "", "Monotributista inscripta en CM"],
    ]
    for i, ej in enumerate(ejemplos, 3):
        for j, val in enumerate(ej, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.fill = PatternFill("solid", fgColor=_GRIS)


def _hoja_jurisdicciones(wb):
    ws = wb.create_sheet("2_Jurisdicciones")
    ws.sheet_properties.tabColor = "2E75B6"

    columnas = [
        ("CUIT *",                 "Mismo CUIT que en hoja 1",                 18),
        ("Código Jurisdicción *",  "3 dígitos: 901, 902, etc.",                18),
        ("Fecha Alta (DD/MM/AAAA)","Ej: 01/04/2012. Dejar vacío si no se sabe",22),
    ]
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 35
    for i, (n, inst, w) in enumerate(columnas, 1):
        _header(ws, 1, i, n)
        _instruccion(ws, 2, i, inst)
        ws.column_dimensions[get_column_letter(i)].width = w

    ejemplos = [
        ["30-71223438-1", "901", "01/04/2012"],
        ["30-71223438-1", "902", "01/04/2012"],
        ["20-12345678-9", "902", "01/01/2020"],
        ["20-12345678-9", "921", "01/01/2020"],
    ]
    for i, ej in enumerate(ejemplos, 3):
        for j, val in enumerate(ej, 1):
            ws.cell(row=i, column=j, value=val).fill = PatternFill("solid", fgColor=_GRIS)


def _hoja_actividades(wb):
    ws = wb.create_sheet("3_Actividades")
    ws.sheet_properties.tabColor = "375623"

    columnas = [
        ("CUIT *",            "Mismo CUIT que en hoja 1",                       18),
        ("Código CUACM *",    "Ej: 266090",                                      15),
        ("Descripción *",     "Descripción completa de la actividad",             45),
        ("Artículo CM *",     "art2 / art6 / art7 / art8 / art9 / art10 / art14",18),
        ("Es Principal",      "SI / NO",                                          12),
        ("Fecha Alta",        "DD/MM/AAAA. Vacío si no se sabe",                  20),
    ]
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 35
    for i, (n, inst, w) in enumerate(columnas, 1):
        _header(ws, 1, i, n)
        _instruccion(ws, 2, i, inst)
        ws.column_dimensions[get_column_letter(i)].width = w

    ejemplos = [
        ["30-71223438-1", "266090", "Fabricación de equipo médico y quirúrgico n.c.p.", "art2", "SI", "01/04/2012"],
        ["20-12345678-9", "731000", "Servicios de publicidad", "art2", "SI", "01/01/2020"],
    ]
    for i, ej in enumerate(ejemplos, 3):
        for j, val in enumerate(ej, 1):
            ws.cell(row=i, column=j, value=val).fill = PatternFill("solid", fgColor=_GRIS)


def _hoja_coeficientes(wb):
    ws = wb.create_sheet("4_Coeficientes")
    ws.sheet_properties.tabColor = "7030A0"

    columnas = [
        ("CUIT *",                "Mismo CUIT que en hoja 1",                        18),
        ("Año *",                 "Ej: 2026",                                         10),
        ("Código Jurisdicción *", "3 dígitos: 901, 902, etc.",                        18),
        ("Coeficiente *",         "4 decimales. Ej: 0.6496  — la suma debe ser 1.0000",22),
    ]
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 50
    for i, (n, inst, w) in enumerate(columnas, 1):
        _header(ws, 1, i, n)
        _instruccion(ws, 2, i, inst)
        ws.column_dimensions[get_column_letter(i)].width = w

    ejemplos = [
        ["30-71223438-1", 2026, "901", 0.6496],
        ["30-71223438-1", 2026, "902", 0.3504],
        ["20-12345678-9", 2026, "902", 0.7000],
        ["20-12345678-9", 2026, "921", 0.3000],
    ]
    for i, ej in enumerate(ejemplos, 3):
        for j, val in enumerate(ej, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.fill = PatternFill("solid", fgColor=_GRIS)
            if j == 4:
                c.number_format = "0.0000"


def _hoja_alicuotas(wb):
    ws = wb.create_sheet("5_Alicuotas")
    ws.sheet_properties.tabColor = "C55A11"

    columnas = [
        ("CUIT *",                "Mismo CUIT que en hoja 1",                 18),
        ("Año *",                 "Ej: 2026",                                  10),
        ("Código Jurisdicción *", "3 dígitos: 901, 902, etc.",                 18),
        ("Código Actividad *",    "Mismo código CUACM que en hoja 3",          16),
        ("Alícuota % *",          "Porcentaje. Ej: 1.50 para 1,50%",           18),
    ]
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 40
    for i, (n, inst, w) in enumerate(columnas, 1):
        _header(ws, 1, i, n)
        _instruccion(ws, 2, i, inst)
        ws.column_dimensions[get_column_letter(i)].width = w

    ejemplos = [
        ["30-71223438-1", 2026, "901", "266090", 1.00],
        ["30-71223438-1", 2026, "902", "266090", 1.50],
        ["20-12345678-9", 2026, "902", "731000", 3.50],
        ["20-12345678-9", 2026, "921", "731000", 3.00],
    ]
    for i, ej in enumerate(ejemplos, 3):
        for j, val in enumerate(ej, 1):
            c = ws.cell(row=i, column=j, value=val)
            c.fill = PatternFill("solid", fgColor=_GRIS)
            if j == 5:
                c.number_format = '0.00"%"'


def _hoja_ayuda(wb):
    ws = wb.create_sheet("Códigos Jurisdicciones")
    ws.sheet_properties.tabColor = "808080"

    _header(ws, 1, 1, "Código", bg="404040")
    _header(ws, 1, 2, "Jurisdicción", bg="404040")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30

    jurs = [
        ("901","Ciudad Autónoma de Buenos Aires"),("902","Buenos Aires"),
        ("903","Catamarca"),("904","Córdoba"),("905","Corrientes"),
        ("906","Chaco"),("907","Chubut"),("908","Entre Ríos"),
        ("909","Formosa"),("910","Jujuy"),("911","La Pampa"),
        ("912","La Rioja"),("913","Mendoza"),("914","Misiones"),
        ("915","Neuquén"),("916","Río Negro"),("917","Salta"),
        ("918","San Juan"),("919","San Luis"),("920","Santa Cruz"),
        ("921","Santa Fe"),("922","Santiago del Estero"),
        ("923","Tucumán"),("924","Tierra del Fuego"),
    ]
    for i, (cod, nom) in enumerate(jurs, 2):
        ws.cell(row=i, column=1, value=cod).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=2, value=nom)
