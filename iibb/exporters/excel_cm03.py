"""
Exportador Excel del papel de trabajo CM03.
Genera un .xlsx con 2 hojas: 'Resumen' y 'Por Jurisdiccion'.
"""
from datetime import datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from iibb.calculo.cm03 import LiquidacionCM03, fmt_money

MESES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

_AZUL = "1F4E79"
_AZUL_CLARO = "BDD7EE"
_VERDE = "375623"
_VERDE_CLARO = "E2EFDA"
_AMARILLO = "FFF2CC"
_GRIS = "F2F2F2"


def _header_style(cell, bg=_AZUL, fg="FFFFFF", bold=True):
    cell.font = Font(bold=bold, color=fg, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _border_thin():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _money_cell(ws, row, col, value: Decimal):
    cell = ws.cell(row=row, column=col, value=float(value))
    cell.number_format = '#,##0.00'
    cell.alignment = Alignment(horizontal="right")
    return cell


def exportar_excel_cm03(liq: LiquidacionCM03, destino: Path | None = None) -> Path:
    """
    Genera el papel de trabajo Excel para una liquidación CM03.
    Si destino es None, guarda en ~/Downloads/.
    Devuelve la ruta del archivo generado.
    """
    if destino is None:
        downloads = Path.home() / "Downloads"
        downloads.mkdir(exist_ok=True)
        nombre = (
            f"CM03_{liq.contribuyente_cuit.replace('-', '')}_{liq.anio}"
            f"{liq.mes:02d}_{datetime.now().strftime('%H%M%S')}.xlsx"
        )
        destino = downloads / nombre

    wb = openpyxl.Workbook()

    _hoja_resumen(wb, liq)
    _hoja_por_jurisdiccion(wb, liq)

    # Eliminar hoja default vacía si quedó
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    wb.save(destino)
    return destino


def _hoja_resumen(wb, liq: LiquidacionCM03):
    ws = wb.create_sheet("Resumen", 0)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 22

    mes_nombre = MESES.get(liq.mes, str(liq.mes))

    # Título
    ws.merge_cells("A1:B1")
    titulo = ws["A1"]
    titulo.value = f"Liquidación CM03 — {mes_nombre} {liq.anio}"
    titulo.font = Font(bold=True, size=14, color=_AZUL)
    titulo.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:B2")
    sub = ws["A2"]
    sub.value = f"{liq.contribuyente_razon}  ({liq.contribuyente_cuit})"
    sub.font = Font(bold=True, size=11)
    sub.alignment = Alignment(horizontal="center")

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18

    # Datos generales
    datos = [
        ("", ""),
        ("Período", f"{mes_nombre} {liq.anio}"),
        ("CUIT", liq.contribuyente_cuit),
        ("Razón Social", liq.contribuyente_razon),
        ("Ingresos Netos del Mes", float(liq.ingresos_totales)),
        ("", ""),
    ]
    row = 3
    for label, val in datos:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=row, column=2, value=val)
        if isinstance(val, float):
            c.number_format = '#,##0.00'
            c.alignment = Alignment(horizontal="right")
        row += 1

    # Tabla por jurisdiccion
    ws.cell(row=row, column=1, value="Jurisdicción").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=_AZUL)
    ws.cell(row=row, column=2, value="Monto a Pagar").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=_AZUL)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    row += 1

    for r in liq.resultados:
        ws.cell(row=row, column=1, value=f"{r.codigo} – {r.nombre}")
        _money_cell(ws, row, 2, r.monto_a_pagar)
        row += 1

    # Total
    ws.cell(row=row, column=1, value="TOTAL A PAGAR").font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=_VERDE)
    c_total = _money_cell(ws, row, 2, liq.monto_total_a_pagar)
    c_total.font = Font(bold=True, size=12, color="FFFFFF")
    c_total.fill = PatternFill("solid", fgColor=_VERDE)


def _hoja_por_jurisdiccion(wb, liq: LiquidacionCM03):
    ws = wb.create_sheet("Por Jurisdicción", 1)

    mes_nombre = MESES.get(liq.mes, str(liq.mes))
    juris_nombres = [f"{r.codigo} – {r.nombre}" for r in liq.resultados]
    n_juris = len(liq.resultados)

    # Columnas: A = concepto, B... = jurisdicciones, última = TOTAL
    ws.column_dimensions["A"].width = 32
    for i in range(n_juris + 1):
        ws.column_dimensions[get_column_letter(i + 2)].width = 22

    row = 1
    # Título
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_juris + 2)
    t = ws.cell(row=row, column=1, value=f"CM03 Detallado — {mes_nombre} {liq.anio} — {liq.contribuyente_razon}")
    t.font = Font(bold=True, size=13, color=_AZUL)
    t.alignment = Alignment(horizontal="center")
    row += 1

    # Headers de columnas
    ws.cell(row=row, column=1, value="Concepto")
    _header_style(ws.cell(row=row, column=1))
    for i, nombre in enumerate(juris_nombres):
        c = ws.cell(row=row, column=i + 2, value=nombre)
        _header_style(c)
    c_tot = ws.cell(row=row, column=n_juris + 2, value="TOTAL")
    _header_style(c_tot)
    row += 1

    conceptos = [
        ("Coeficiente",            [r.coeficiente for r in liq.resultados],            None,  "coef"),
        ("Alícuota",               [r.alicuota for r in liq.resultados],                None,  "porc"),
        ("Base Imponible",         [r.base_imponible for r in liq.resultados],          True,  "money"),
        ("Impuesto Determinado",   [r.impuesto_determinado for r in liq.resultados],    True,  "money"),
        ("(−) SAF Período Anterior",[r.saf_anterior for r in liq.resultados],           True,  "money"),
        ("(−) Retenciones",        [r.retenciones for r in liq.resultados],             True,  "money"),
        ("(−) Percepciones",       [r.percepciones for r in liq.resultados],            True,  "money"),
        ("(−) SIRCREB",            [r.sircreb for r in liq.resultados],                 True,  "money"),
        ("(−) Perc. Aduaneras",    [r.perc_aduaneras for r in liq.resultados],          True,  "money"),
        ("(−) Pagos a Cuenta",     [r.pagos_a_cuenta for r in liq.resultados],          True,  "money"),
        ("Total Deducciones",      [r.total_deducciones for r in liq.resultados],       True,  "money"),
        ("MONTO A PAGAR",          [r.monto_a_pagar for r in liq.resultados],           True,  "pagar"),
        ("Saldo a Favor",          [r.saldo_a_favor for r in liq.resultados],           True,  "money"),
    ]

    for label, valores, sumable, tipo in conceptos:
        is_pagar = tipo == "pagar"
        bg = _VERDE_CLARO if is_pagar else (_GRIS if row % 2 == 0 else "FFFFFF")
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(bold=is_pagar)
        label_cell.fill = PatternFill("solid", fgColor=bg)
        label_cell.border = _border_thin()

        total_val = Decimal("0")
        for i, val in enumerate(valores):
            c = ws.cell(row=row, column=i + 2)
            c.fill = PatternFill("solid", fgColor=bg)
            c.border = _border_thin()
            if tipo == "coef":
                c.value = float(val)
                c.number_format = "0.0000"
                c.alignment = Alignment(horizontal="center")
            elif tipo == "porc":
                c.value = float(val) * 100
                c.number_format = '0.00"%"'
                c.alignment = Alignment(horizontal="center")
            else:
                c.value = float(val)
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right")
                c.font = Font(bold=is_pagar)
            if sumable:
                total_val += Decimal(str(val))

        # Columna TOTAL
        c_t = ws.cell(row=row, column=n_juris + 2)
        c_t.fill = PatternFill("solid", fgColor=_AZUL_CLARO if is_pagar else bg)
        c_t.border = _border_thin()
        if sumable:
            c_t.value = float(total_val)
            c_t.number_format = '#,##0.00'
            c_t.alignment = Alignment(horizontal="right")
            c_t.font = Font(bold=is_pagar)
        else:
            c_t.value = "—"
            c_t.alignment = Alignment(horizontal="center")

        row += 1
