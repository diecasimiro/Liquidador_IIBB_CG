"""
Genera el Excel de Resumen Anual CM03 para un contribuyente.
"""
from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from iibb.service.resumen_anual import ResumenAnual

MESES = {
    1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
    7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre",
}

_AZUL    = "1F4E79"
_AZUL_CL = "BDD7EE"
_VERDE   = "375623"
_GRIS    = "F2F2F2"
_BLANCO  = "FFFFFF"

_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

COLUMNAS = [
    ("Base Imponible",         "base_imponible"),
    ("Impuesto Determinado",   "impuesto_determinado"),
    ("Retenciones",            "retenciones"),
    ("Percepciones",           "percepciones"),
    ("Perc. Aduaneras",        "perc_aduaneras"),
    ("SIRCREB",                "sircreb"),
    ("Saldo a Favor (Dic.)",   "saldo_a_favor_dic"),
    ("IIBB a Pagar",           "iibb_a_pagar"),
]


def _celda(ws, row, col, value, bold=False, bg=None, fg="000000",
           num_format=None, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fg)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    c.border = _BORDER
    if num_format:
        c.number_format = num_format
    return c


def exportar_excel_resumen_anual(
    resumen: ResumenAnual,
    razon_social: str,
    cuit: str,
    destino: Path | None = None,
) -> Path:
    if destino is None:
        downloads = Path.home() / "Downloads"
        downloads.mkdir(exist_ok=True)
        destino = downloads / f"resumen_anual_{cuit}_{resumen.anio}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Resumen {resumen.anio}"

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(COLUMNAS))
    c = ws.cell(row=1, column=1, value=f"RESUMEN ANUAL IIBB CM03 — {resumen.anio}")
    c.font = Font(bold=True, size=14, color=_BLANCO)
    c.fill = PatternFill("solid", fgColor=_AZUL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=1 + len(COLUMNAS))
    c2 = ws.cell(row=2, column=1, value=f"{razon_social}   CUIT: {cuit}")
    c2.font = Font(bold=True, size=11)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    meses_str = ", ".join(MESES.get(m, str(m)) for m in resumen.meses_liquidados)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=1 + len(COLUMNAS))
    c3 = ws.cell(row=3, column=1,
                 value=f"Meses liquidados: {meses_str or 'ninguno'}")
    c3.font = Font(italic=True, size=9)
    c3.alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 16

    # ── Encabezados ───────────────────────────────────────────────────────────
    ROW_HDR = 4
    _celda(ws, ROW_HDR, 1, "Jurisdicción", bold=True, bg=_AZUL, fg=_BLANCO, align="center")
    ws.column_dimensions["A"].width = 28
    for i, (label, _) in enumerate(COLUMNAS, 2):
        _celda(ws, ROW_HDR, i, label, bold=True, bg=_AZUL, fg=_BLANCO, align="center")
        ws.column_dimensions[get_column_letter(i)].width = 20
    ws.row_dimensions[ROW_HDR].height = 36

    # ── Datos ─────────────────────────────────────────────────────────────────
    FMT = '#,##0.00'
    totales = {attr: 0 for _, attr in COLUMNAS}

    for idx, fila in enumerate(resumen.filas):
        row = ROW_HDR + 1 + idx
        bg = _GRIS if idx % 2 == 0 else _BLANCO
        _celda(ws, row, 1, f"{fila.codigo} — {fila.nombre}", bg=bg)
        for col_i, (_, attr) in enumerate(COLUMNAS, 2):
            val = float(getattr(fila, attr))
            _celda(ws, row, col_i, val, bg=bg, num_format=FMT, align="right")
            totales[attr] += val

    # ── Fila totales ──────────────────────────────────────────────────────────
    row_tot = ROW_HDR + 1 + len(resumen.filas)
    _celda(ws, row_tot, 1, "TOTAL", bold=True, bg=_VERDE, fg=_BLANCO)
    for col_i, (_, attr) in enumerate(COLUMNAS, 2):
        _celda(ws, row_tot, col_i, totales[attr],
               bold=True, bg=_VERDE, fg=_BLANCO, num_format=FMT, align="right")
    ws.row_dimensions[row_tot].height = 18

    # ── Nota al pie ───────────────────────────────────────────────────────────
    row_nota = row_tot + 2
    ws.merge_cells(start_row=row_nota, start_column=1,
                   end_row=row_nota, end_column=1 + len(COLUMNAS))
    nota = ws.cell(row=row_nota, column=1,
                   value="(*) Saldo a Favor: corresponde al mes de Diciembre. "
                         "Si no está liquidado figura en $0,00.")
    nota.font = Font(italic=True, size=8)

    wb.save(destino)
    return destino
