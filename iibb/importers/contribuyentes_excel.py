"""
Importador masivo de contribuyentes desde la plantilla Excel de 5 hojas.
Devuelve un resumen con filas importadas, advertencias y errores.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from sqlalchemy.orm import Session

from iibb.service.contribuyente import (
    crear_contribuyente,
    obtener_por_cuit,
    validar_cuit,
    agregar_jurisdiccion,
    guardar_coeficientes_anio,
    guardar_alicuotas_anio,
    actualizar_datos_basicos,
)
from iibb.models.contribuyente import ActividadInscripta


@dataclass
class ResultadoImportacion:
    contribuyentes_nuevos: int = 0
    contribuyentes_actualizados: int = 0
    errores: list[str] = field(default_factory=list)
    advertencias: list[str] = field(default_factory=list)


# ─── helpers ──────────────────────────────────────────────────────────────────

def _str(v) -> str:
    return str(v).strip() if v is not None else ""


def _fecha(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, (date, datetime)):
        return v.date() if isinstance(v, datetime) else v
    s = _str(v)
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _decimal(v) -> Decimal | None:
    if v is None:
        return None
    try:
        return Decimal(str(v).replace(",", ".").strip())
    except InvalidOperation:
        return None


def _rows(ws, desde: int = 3):
    """Itera filas de la hoja desde `desde` en adelante, omite filas vacías."""
    for row in ws.iter_rows(min_row=desde, values_only=True):
        if any(c is not None and _str(c) != "" for c in row):
            yield row


# ─── lector por hoja ──────────────────────────────────────────────────────────

def _leer_contribuyentes(ws) -> dict[str, dict]:
    """Devuelve dict cuit→datos_básicos."""
    datos: dict[str, dict] = {}
    for row in _rows(ws):
        cuit_raw = _str(row[0] if len(row) > 0 else "")
        if not cuit_raw:
            continue
        ok, cuit = validar_cuit(cuit_raw)
        if not ok:
            continue  # se reporta luego
        datos[cuit] = {
            "cuit": cuit,
            "razon_social": _str(row[1] if len(row) > 1 else ""),
            "nro_inscripcion_cm": _str(row[2] if len(row) > 2 else "") or None,
            "jurisdiccion_sede": _str(row[3] if len(row) > 3 else "") or None,
            "tipo_contribuyente": _str(row[4] if len(row) > 4 else "") or None,
            "naturaleza_juridica": _str(row[5] if len(row) > 5 else "") or None,
            "cierre_ejercicio_mes": int(row[6]) if row[6] is not None and str(row[6]).strip().isdigit() else None,
            "cierre_ejercicio_dia": int(row[7]) if row[7] is not None and str(row[7]).strip().isdigit() else None,
            "domicilio_fiscal": _str(row[8] if len(row) > 8 else "") or None,
            "domicilio_actividades": _str(row[9] if len(row) > 9 else "") or None,
            "notas": _str(row[10] if len(row) > 10 else "") or None,
        }
    return datos


def _leer_jurisdicciones(ws) -> dict[str, list[dict]]:
    jurs: dict[str, list[dict]] = {}
    for row in _rows(ws):
        cuit_raw = _str(row[0] if len(row) > 0 else "")
        if not cuit_raw:
            continue
        ok, cuit = validar_cuit(cuit_raw)
        if not ok:
            continue
        codigo = _str(row[1] if len(row) > 1 else "")
        if not codigo:
            continue
        jurs.setdefault(cuit, []).append({
            "codigo": codigo,
            "fecha_alta": _fecha(row[2] if len(row) > 2 else None),
        })
    return jurs


def _leer_actividades(ws) -> dict[str, list[dict]]:
    acts: dict[str, list[dict]] = {}
    for row in _rows(ws):
        cuit_raw = _str(row[0] if len(row) > 0 else "")
        if not cuit_raw:
            continue
        ok, cuit = validar_cuit(cuit_raw)
        if not ok:
            continue
        codigo = _str(row[1] if len(row) > 1 else "")
        descripcion = _str(row[2] if len(row) > 2 else "")
        if not codigo or not descripcion:
            continue
        es_principal_raw = _str(row[4] if len(row) > 4 else "").upper()
        acts.setdefault(cuit, []).append({
            "codigo_cuacm": codigo,
            "descripcion": descripcion,
            "articulo_cm": _str(row[3] if len(row) > 3 else "") or "art2",
            "es_principal": es_principal_raw in ("SI", "SÍ", "S", "YES", "TRUE", "1"),
            "fecha_alta": _fecha(row[5] if len(row) > 5 else None),
        })
    return acts


def _leer_coeficientes(ws) -> dict[str, list[dict]]:
    coefs: dict[str, list[dict]] = {}
    for row in _rows(ws):
        cuit_raw = _str(row[0] if len(row) > 0 else "")
        if not cuit_raw:
            continue
        ok, cuit = validar_cuit(cuit_raw)
        if not ok:
            continue
        anio_raw = row[1] if len(row) > 1 else None
        codigo = _str(row[2] if len(row) > 2 else "")
        valor = _decimal(row[3] if len(row) > 3 else None)
        if not anio_raw or not codigo or valor is None:
            continue
        try:
            anio = int(anio_raw)
        except (ValueError, TypeError):
            continue
        coefs.setdefault(cuit, []).append({
            "anio": anio,
            "jurisdiccion_codigo": codigo,
            "valor": valor,
        })
    return coefs


def _leer_alicuotas(ws) -> dict[str, list[dict]]:
    alis: dict[str, list[dict]] = {}
    for row in _rows(ws):
        cuit_raw = _str(row[0] if len(row) > 0 else "")
        if not cuit_raw:
            continue
        ok, cuit = validar_cuit(cuit_raw)
        if not ok:
            continue
        anio_raw = row[1] if len(row) > 1 else None
        codigo_jur = _str(row[2] if len(row) > 2 else "")
        codigo_act = _str(row[3] if len(row) > 3 else "")
        valor = _decimal(row[4] if len(row) > 4 else None)
        if not anio_raw or not codigo_jur or not codigo_act or valor is None:
            continue
        try:
            anio = int(anio_raw)
        except (ValueError, TypeError):
            continue
        alis.setdefault(cuit, []).append({
            "anio": anio,
            "jurisdiccion_codigo": codigo_jur,
            "codigo_actividad": codigo_act,
            "valor": valor / Decimal("100"),  # plantilla en %, guardamos fracción
        })
    return alis


# ─── función principal ────────────────────────────────────────────────────────

def importar_contribuyentes_excel(
    archivo: Path | str,
    session: Session,
    tenant_id: int = 1,
) -> ResultadoImportacion:
    resultado = ResultadoImportacion()

    try:
        wb = openpyxl.load_workbook(str(archivo), data_only=True)
    except Exception as e:
        resultado.errores.append(f"No se pudo abrir el archivo: {e}")
        return resultado

    # Hojas por nombre (flexible)
    def _hoja(parcial: str):
        for name in wb.sheetnames:
            if parcial.lower() in name.lower():
                return wb[name]
        return None

    ws_cont = _hoja("contribuyente")
    ws_jur  = _hoja("jurisdicc")
    ws_act  = _hoja("actividad")
    ws_coe  = _hoja("coeficiente")
    ws_ali  = _hoja("alicuota") or _hoja("alícuota")

    if ws_cont is None:
        resultado.errores.append("No se encontró la hoja '1_Contribuyentes'.")
        return resultado

    contribuyentes = _leer_contribuyentes(ws_cont)
    jurisdicciones = _leer_jurisdicciones(ws_jur) if ws_jur else {}
    actividades    = _leer_actividades(ws_act)    if ws_act  else {}
    coeficientes   = _leer_coeficientes(ws_coe)   if ws_coe  else {}
    alicuotas      = _leer_alicuotas(ws_ali)      if ws_ali  else {}

    # Validar CUITs de la hoja 1 antes de tocar la BD
    for row in _rows(ws_cont):
        cuit_raw = _str(row[0] if len(row) > 0 else "")
        if not cuit_raw:
            continue
        ok, _ = validar_cuit(cuit_raw)
        if not ok:
            resultado.errores.append(f"CUIT inválido en hoja Contribuyentes: '{cuit_raw}'")

    if not contribuyentes and not resultado.errores:
        resultado.advertencias.append("La hoja de contribuyentes no tiene filas de datos.")
        return resultado

    for cuit, datos in contribuyentes.items():
        if not datos["razon_social"]:
            resultado.errores.append(f"CUIT {cuit}: falta Razón Social.")
            continue

        jurs_contrib = jurisdicciones.get(cuit, [])
        acts_contrib = actividades.get(cuit, [])
        coefs_contrib = coeficientes.get(cuit, [])
        alis_contrib  = alicuotas.get(cuit, [])

        existente = obtener_por_cuit(session, cuit)

        try:
            if existente is None:
                # Crear nuevo
                crear_contribuyente(
                    session,
                    cuit=cuit,
                    razon_social=datos["razon_social"],
                    nro_inscripcion_cm=datos["nro_inscripcion_cm"],
                    jurisdiccion_sede=datos["jurisdiccion_sede"],
                    tipo_contribuyente=datos["tipo_contribuyente"],
                    naturaleza_juridica=datos["naturaleza_juridica"],
                    cierre_ejercicio_mes=datos["cierre_ejercicio_mes"],
                    cierre_ejercicio_dia=datos["cierre_ejercicio_dia"],
                    domicilio_fiscal=datos["domicilio_fiscal"],
                    domicilio_actividades=datos["domicilio_actividades"],
                    notas=datos["notas"],
                    tenant_id=tenant_id,
                    jurisdicciones=jurs_contrib,
                    actividades=acts_contrib,
                    coeficientes=coefs_contrib,
                    alicuotas=alis_contrib,
                )
                resultado.contribuyentes_nuevos += 1
            else:
                # Actualizar datos básicos
                actualizar_datos_basicos(
                    session,
                    existente.id,
                    razon_social=datos["razon_social"].strip().upper(),
                    nro_inscripcion_cm=datos["nro_inscripcion_cm"],
                    jurisdiccion_sede=datos["jurisdiccion_sede"],
                    tipo_contribuyente=datos["tipo_contribuyente"],
                    naturaleza_juridica=datos["naturaleza_juridica"],
                    cierre_ejercicio_mes=datos["cierre_ejercicio_mes"],
                    cierre_ejercicio_dia=datos["cierre_ejercicio_dia"],
                    domicilio_fiscal=datos["domicilio_fiscal"],
                    domicilio_actividades=datos["domicilio_actividades"],
                    notas=datos["notas"],
                )
                # Agregar/reactivar jurisdicciones
                for j in jurs_contrib:
                    agregar_jurisdiccion(
                        session,
                        existente.id,
                        j["codigo"],
                        j.get("fecha_alta"),
                        tenant_id=tenant_id,
                    )
                # Agregar actividades que no existan
                for a in acts_contrib:
                    ya = (
                        session.query(ActividadInscripta)
                        .filter_by(contribuyente_id=existente.id, codigo_cuacm=a["codigo_cuacm"])
                        .first()
                    )
                    if ya is None:
                        session.add(ActividadInscripta(
                            tenant_id=tenant_id,
                            contribuyente_id=existente.id,
                            codigo_cuacm=a["codigo_cuacm"],
                            descripcion=a["descripcion"],
                            articulo_cm=a.get("articulo_cm", "art2"),
                            es_principal=a.get("es_principal", False),
                            fecha_alta=a.get("fecha_alta"),
                            activa=True,
                        ))
                # Coeficientes: reemplazar por año
                anios_coef = {c["anio"] for c in coefs_contrib}
                for anio in anios_coef:
                    guardar_coeficientes_anio(
                        session,
                        existente.id,
                        anio,
                        [c for c in coefs_contrib if c["anio"] == anio],
                        tenant_id=tenant_id,
                    )
                # Alícuotas: reemplazar por año
                anios_ali = {a["anio"] for a in alis_contrib}
                for anio in anios_ali:
                    guardar_alicuotas_anio(
                        session,
                        existente.id,
                        anio,
                        [a for a in alis_contrib if a["anio"] == anio],
                        tenant_id=tenant_id,
                    )
                resultado.contribuyentes_actualizados += 1
        except Exception as e:
            resultado.errores.append(f"CUIT {cuit}: {e}")

    return resultado
